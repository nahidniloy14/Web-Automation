import openpyxl


class HomePageData:
    data_tuple = [("Nahid Hassan", "Niloy", "1414"),
                  ("Nabil", "Hossain", "0417"),
                  ("Joy", "Rahaman", "0457")]
    data_dictonary = [{"firstName": "Nahid Hassan", "lastName": "Niloy", "postalCode": "1414"},
                      {"firstName": "Nabil", "lastName": "Hossain", "postalCode": "1424"}, ]

    # get value from excel
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Owner\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

