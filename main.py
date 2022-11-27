import openpyxl
book=openpyxl.load_workbook("C:\\Users\\lm\\OneDrive - American International University-Bangladesh\\Documents\\RoadToSqa.xlsx")#file path
sheet=book.active #active excel sheet
cell=sheet.cell(row=15,column=2)
print(cell.value)
sheet.cell(row=1,column=5).value="Experiment"
print(sheet.cell(row=1,column=5).value)#value will be added to the sheet
print(sheet.max_row)
print(sheet.max_column)
