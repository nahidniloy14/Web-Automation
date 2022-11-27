import openpyxl

book = openpyxl.load_workbook(
    "C:\\Users\\lm\\OneDrive - American International University-Bangladesh\\Documents\\RoadToSqa.xlsx")  # file path
sheet = book.active  # active excel sheet
sheet.cell(row=2, column=5).value = "Experiment"
print(sheet.cell(row=2, column=5).value)  # value will be added to the sheet
