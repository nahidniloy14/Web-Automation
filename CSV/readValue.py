import openpyxl

book = openpyxl.load_workbook(
    "C:\\Users\\lm\\OneDrive - American International University-Bangladesh\\Documents\\RoadToSqa.xlsx")  # file path
sheet = book.active  # active Excel sheet
cell = sheet.cell(row=15, column=2)
print(cell.value)  # 2022-05-25 0
print(sheet.max_row)  # 101
print(sheet.max_column)  # 4
print(sheet['A9'].value)  # 8
# +1 to prevent excluding the last value
Dict = {}
for i in range(1, sheet.max_row + 1):  # to get rows
    for j in range(2, sheet.max_column + 1):  # to get columns
        print(sheet.cell(row=i, column=j).value)


# inject value in excel dictionary
# for i in range(1, sheet.max_row + 1):  # to get rows
#     if sheet.cell(row=i, column=1).value == "Topic":
#
#         for j in range(2, sheet.max_column + 1):  # to get columns
#             #Dict["lastname"] = "shetty"
#             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
#
# print(Dict)
